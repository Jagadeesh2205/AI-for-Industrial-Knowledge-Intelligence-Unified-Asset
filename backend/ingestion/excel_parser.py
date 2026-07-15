"""
Excel/CSV Parser — Converts spreadsheet data to structured text for indexing.

Handles both .xlsx (openpyxl) and .csv (built-in csv module).
Converts each sheet/table to markdown for LLM consumption.
"""

import csv
from pathlib import Path
from typing import Optional


class ExcelParser:
    """Parse Excel and CSV files into structured text."""

    def parse(self, filepath: str) -> dict:
        """
        Parse an Excel or CSV file.
        
        Returns same structure as PDFParser for consistency.
        """
        ext = Path(filepath).suffix.lower()
        
        if ext == ".csv":
            return self._parse_csv(filepath)
        else:
            return self._parse_xlsx(filepath)

    def _parse_xlsx(self, filepath: str) -> dict:
        """Parse an Excel file using openpyxl."""
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(filepath, read_only=True, data_only=True)
            pages = []
            full_text_parts = []
            
            for sheet_idx, sheet_name in enumerate(wb.sheetnames):
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                
                if not rows:
                    continue
                
                # Convert to markdown table
                md_table = self._rows_to_markdown(rows)
                text = f"## Sheet: {sheet_name}\n\n{md_table}"
                
                pages.append({
                    "page_num": sheet_idx + 1,
                    "text": text,
                    "tables": [md_table],
                    "section_header": {"number": "", "title": sheet_name},
                    "has_images": False,
                })
                full_text_parts.append(text)
            
            wb.close()
            
            return {
                "pages": pages,
                "metadata": {
                    "title": Path(filepath).stem,
                    "author": "Unknown",
                    "page_count": len(pages),
                    "file_size": Path(filepath).stat().st_size,
                },
                "full_text": "\n\n".join(full_text_parts),
                "sections": [{"number": "", "title": name, "page": i+1} for i, name in enumerate(wb.sheetnames)],
            }
        except ImportError:
            # Fallback if openpyxl not available
            return {
                "pages": [{"page_num": 1, "text": f"[Excel file: {Path(filepath).name}]", "tables": [], "section_header": None, "has_images": False}],
                "metadata": {"title": Path(filepath).stem, "author": "Unknown", "page_count": 1, "file_size": Path(filepath).stat().st_size},
                "full_text": f"[Excel file: {Path(filepath).name}]",
                "sections": [],
            }

    def _parse_csv(self, filepath: str) -> dict:
        """Parse a CSV file."""
        with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
            reader = csv.reader(f)
            rows = [row for row in reader]
        
        if not rows:
            return {
                "pages": [],
                "metadata": {"title": Path(filepath).stem, "author": "Unknown", "page_count": 0, "file_size": 0},
                "full_text": "",
                "sections": [],
            }
        
        md_table = self._rows_to_markdown(rows)
        
        return {
            "pages": [{"page_num": 1, "text": md_table, "tables": [md_table], "section_header": None, "has_images": False}],
            "metadata": {
                "title": Path(filepath).stem,
                "author": "Unknown",
                "page_count": 1,
                "file_size": Path(filepath).stat().st_size,
            },
            "full_text": md_table,
            "sections": [],
        }

    def _rows_to_markdown(self, rows: list) -> str:
        """Convert a list of rows to a markdown table."""
        if not rows:
            return ""
        
        # Clean cells
        cleaned = []
        for row in rows:
            cleaned_row = [str(cell).strip().replace("\n", " ") if cell is not None else "" for cell in row]
            cleaned.append(cleaned_row)
        
        # Determine max columns
        max_cols = max(len(row) for row in cleaned) if cleaned else 0
        if max_cols == 0:
            return ""
        
        # Pad rows
        for row in cleaned:
            while len(row) < max_cols:
                row.append("")
        
        # Build markdown
        header = "| " + " | ".join(cleaned[0]) + " |"
        separator = "| " + " | ".join(["---"] * max_cols) + " |"
        body_rows = ["| " + " | ".join(row) + " |" for row in cleaned[1:]]
        
        return "\n".join([header, separator] + body_rows)
